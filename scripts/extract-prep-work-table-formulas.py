#!/usr/bin/env python3
"""Extract Prep Work table formulas + values to prep-work-table-cells.json."""
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

path = Path(
    sys.argv[1]
    if len(sys.argv) > 1
    else r"c:\Users\b.carson\Downloads\New Zealand v South Africa 63406779 (1).xlsm"
)
out_path = Path(__file__).resolve().parent.parent / "lib" / "workbooks" / "prep-work-table-cells.json"

wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
wbv = openpyxl.load_workbook(path, read_only=True, data_only=True)
prep = wb["Prep Work"]
prepv = wbv["Prep Work"]


def dump_range(min_row, max_row, min_col, max_col):
    cells = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            addr = f"{get_column_letter(c)}{r}"
            f = prep[addr].value
            v = prepv[addr].value
            if f is None and v is None:
                continue
            formula = None
            if f is not None:
                formula = str(f)
            cells.append(
                {
                    "address": addr,
                    "row": r,
                    "col": get_column_letter(c),
                    "formula": formula,
                    "value": v,
                }
            )
    return cells


out = {
    "fixtureId": "nz-sa-63406779",
    "label": "New Zealand v South Africa (63406779)",
    "sheet": "Prep Work",
    "tables": [
        {
            "id": "table-1",
            "name": "Table 1 — Per-innings historical blends",
            "range": "K2:P18",
            "cells": dump_range(2, 18, 11, 16),
        },
        {
            "id": "table-2",
            "name": "Table 2 — Match-level model stats",
            "range": "T2:AB10",
            "cells": dump_range(2, 10, 20, 28),
        },
    ],
}

out_path.write_text(json.dumps(out, indent=2, default=str), encoding="utf-8")
print(f"wrote {out_path} ({sum(len(t['cells']) for t in out['tables'])} cells)")
wb.close()
wbv.close()
