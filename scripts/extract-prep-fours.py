#!/usr/bin/env python3
"""Extract player fours/sixes columns from Prep Work."""
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import coordinate_to_tuple, get_column_letter

path = Path(
    sys.argv[1]
    if len(sys.argv) > 1
    else r"c:\Users\b.carson\Downloads\New Zealand v South Africa 63406779 (1).xlsm"
)

wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
prep = wb["Prep Work"]


def cell(addr: str):
    row, col = coordinate_to_tuple(addr)
    for r in prep.iter_rows(
        min_row=row, max_row=row, min_col=col, max_col=col, values_only=True
    ):
        return r[0]
    return None


# Header row for player stats area (around row 23-35)
for row in range(22, 38):
    vals = []
    for col in range(12, 18):  # L through Q
        addr = f"{get_column_letter(col)}{row}"
        v = cell(addr)
        if v is not None:
            vals.append(f"{addr}={v}")
    if vals:
        print(f"row {row}: " + ", ".join(vals))

print("\n--- Sample player rows 36-38 cols N-P ---")
for row in range(24, 40):
    n, o, p = cell(f"N{row}"), cell(f"O{row}"), cell(f"P{row}")
    if any(x is not None for x in (n, o, p)):
        print(f"row {row}: N={n}, O={o}, P={p}")

print("\n--- Sum O24:O34 (home XI fours?) ---")
total = 0
for row in range(24, 35):
    v = cell(f"O{row}")
    if isinstance(v, (int, float)):
        total += v
print(f"sum O24:O34 = {total}")

print("\n--- Sum P24:P34 (home XI sixes?) ---")
total6 = 0
for row in range(24, 35):
    v = cell(f"P{row}")
    if isinstance(v, (int, float)):
        total6 += v
print(f"sum P24:P34 = {total6}")

# Away team might be further down - scan O column
print("\n--- All numeric O column rows 24-60 ---")
for row in range(24, 61):
    v = cell(f"O{row}")
    if isinstance(v, (int, float)) and v > 0:
        name = cell(f"B{row}") or cell(f"G{row}")
        print(f"O{row}={v} ({name})")

wb.close()
