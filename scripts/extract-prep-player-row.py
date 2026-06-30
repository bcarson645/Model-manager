#!/usr/bin/env python3
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

path = Path(
    sys.argv[1]
    if len(sys.argv) > 1
    else r"c:\Users\b.carson\Downloads\New Zealand v South Africa 63406779 (1).xlsm"
)
out_path = (
    Path(__file__).resolve().parent.parent
    / "lib"
    / "workbooks"
    / "prep-player-row-cells.json"
)

wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
wbv = openpyxl.load_workbook(path, read_only=True, data_only=True)
prep = wb["Prep Work"]
prepv = wbv["Prep Work"]


def formula_text(value):
    if value is None:
        return None
    if isinstance(value, ArrayFormula):
        return value.text
    return str(value)


def dump_range(min_row, max_row, min_col, max_col):
    cells = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            addr = f"{get_column_letter(c)}{r}"
            f = prep[addr].value
            v = prepv[addr].value
            if f is None and v is None:
                continue
            cells.append(
                {
                    "address": addr,
                    "row": r,
                    "col": get_column_letter(c),
                    "formula": formula_text(f),
                    "value": v,
                }
            )
    return cells


def dump_addrs(addrs):
    cells = []
    for addr in addrs:
        f = prep[addr].value
        v = prepv[addr].value
        if f is None and v is None:
            continue
        row = int("".join(ch for ch in addr if ch.isdigit()))
        col = "".join(ch for ch in addr if ch.isalpha())
        cells.append(
            {
                "address": addr,
                "row": row,
                "col": col,
                "formula": formula_text(f),
                "value": v,
            }
        )
    return cells


def dump_dep_rows(rows):
    cells = []
    for r in rows:
        for c in DEP_COLS:
            addr = f"{get_column_letter(c)}{r}"
            f = prep[addr].value
            v = prepv[addr].value
            if f is None and v is None:
                continue
            cells.append(
                {
                    "address": addr,
                    "row": r,
                    "col": get_column_letter(c),
                    "formula": formula_text(f),
                    "value": v,
                }
            )
    return cells


def player_meta(row):
    name = prepv[f"K{row}"].value
    bat = prepv[f"I{row}"].value
    bowl = prepv[f"J{row}"].value
    pid = prepv[f"H{row}"].value
    if not any(x is not None for x in (name, bat, bowl, pid)):
        return None
    return {
        "row": row,
        "name": name,
        "battingPosition": bat,
        "bowlingPosition": bowl,
        "playerId": pid,
    }


home_rows = list(range(24, 35))
away_rows = list(range(45, 56))
# Columns used upstream of M / O / P on each player row
DEP_COLS = [33, 39, 40, 41, 42, 43, 44, 82, 83, 84, 110, 112, 126]  # AG, AM, AN, AO, AP, AQ, AR, CD, CE, CF, DF, DH, DV
SHARED_ADDRS = ["D3", "D5", "BW3", "M21", "O21", "P21", "BT3", "W63", "Y63", "X63"]

out = {
    "fixtureId": "nz-sa-63406779",
    "label": "New Zealand v South Africa (63406779)",
    "sheet": "Prep Work",
    "headers": dump_range(23, 23, 8, 16),  # H23:P23
    "bowlingHeaders": dump_range(23, 23, 22, 27),  # V23:AA23 (through Z; AA=style)
    "ratingHeaders": dump_range(23, 23, 17, 17),  # Q23
    "sharedInputs": dump_addrs(SHARED_ADDRS) + dump_range(21, 21, 13, 16),
    "squads": [
        {
            "id": "home",
            "name": "Home XI (Team 1)",
            "playerRows": home_rows,
            "teamFoursTotal": {"address": "O36", "formula": formula_text(prep["O36"].value), "value": prepv["O36"].value},
            "teamSixesTotal": {"address": "P36", "formula": formula_text(prep["P36"].value), "value": prepv["P36"].value},
            "cells": dump_range(24, 34, 8, 17),  # H24:Q34 (includes batting rating Q)
            "bowlingCells": dump_range(24, 34, 22, 27),  # V24:AA34
            "dependencyCells": dump_dep_rows(home_rows),
            "players": [m for r in home_rows if (m := player_meta(r))],
        },
        {
            "id": "away",
            "name": "Away XI (Team 2)",
            "playerRows": away_rows,
            "teamFoursTotal": {"address": "O57", "formula": formula_text(prep["O57"].value), "value": prepv["O57"].value},
            "teamSixesTotal": {"address": "P57", "formula": formula_text(prep["P57"].value), "value": prepv["P57"].value},
            "cells": dump_range(45, 55, 8, 17),  # H45:Q55
            "bowlingCells": dump_range(45, 55, 22, 27),
            "dependencyCells": dump_dep_rows(away_rows),
            "players": [m for r in away_rows if (m := player_meta(r))],
        },
    ],
}

out_path.write_text(json.dumps(out, indent=2, default=str), encoding="utf-8")
print(f"wrote {out_path} ({sum(len(s['cells']) for s in out['squads'])} cells)")
wb.close()
wbv.close()
