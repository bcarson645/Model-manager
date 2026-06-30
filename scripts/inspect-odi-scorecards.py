#!/usr/bin/env python3
"""Profile an ODI scorecard Excel export: columns, game grouping, player IDs."""

import json
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

HEADER_ROW = 16
DATA_SHEET = "Data"
DEFAULT_OUT = (
    Path(__file__).resolve().parent.parent / "lib" / "odi-scorecards" / "schema-profile.json"
)


def serialize(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def guess_key_columns(headers: list[str]) -> dict[str, str | None]:
    lower = {h: (h or "").strip().lower() for h in headers}

    def pick(*candidates: str) -> str | None:
        for h, norm in lower.items():
            if any(c in norm for c in candidates):
                return h
        return None

    return {
        "gameId": pick("game id", "game_id", "match id", "match_id", "fixture", "gameid"),
        "date": pick("date", "match date", "start date"),
        "playerId": pick("player id", "player_id", "playerid", "sr player"),
        "playerName": pick("player name", "player_name", "name", "batter", "bowler"),
        "team": pick("team", "side", "squad"),
        "innings": pick("innings", "inning"),
    }


def profile_workbook(path: Path, sheet_name: str | None = None) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb[sheet_name] if sheet_name else wb.active

    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError(f"Sheet {sheet.title!r} is empty")

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
    keys = guess_key_columns(headers)
    col_index = {h: i for i, h in enumerate(headers)}

    records: list[dict] = []
    null_counts = Counter()
    type_samples: dict[str, set[str]] = defaultdict(set)

    for row in rows_iter:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        rec = {}
        for i, h in enumerate(headers):
            v = row[i] if i < len(row) else None
            rec[h] = serialize(v)
            if v is None or str(v).strip() == "":
                null_counts[h] += 1
            else:
                type_samples[h].add(type(v).__name__)
        records.append(rec)

    game_col = keys["gameId"]
    games: dict[str, list[dict]] = defaultdict(list)
    if game_col and game_col in col_index:
        for rec in records:
            gid = rec.get(game_col)
            if gid is not None:
                games[str(gid)].append(rec)

    sample_game_id = next(iter(games), None)
    sample_rows = games[sample_game_id][:6] if sample_game_id else records[:6]

    player_col = keys["playerId"]
    players_per_game: dict[str, int] = {}
    if game_col and player_col:
        for gid, game_rows in games.items():
            players = {r.get(player_col) for r in game_rows if r.get(player_col) is not None}
            players_per_game[gid] = len(players)

    row_counts = list(players_per_game.values()) if players_per_game else []
    row_count_dist = Counter(row_counts)

    profile = {
        "sourceFile": path.name,
        "sheetName": sheet.title,
        "rowCount": len(records),
        "columnCount": len(headers),
        "headers": headers,
        "inferredKeys": keys,
        "columnStats": [
            {
                "header": h,
                "nullCount": null_counts[h],
                "nonNullCount": len(records) - null_counts[h],
                "valueTypes": sorted(type_samples[h]),
                "sampleValues": [
                    records[i][h]
                    for i in range(min(5, len(records)))
                    if records[i].get(h) is not None
                ],
            }
            for h in headers
        ],
        "gameStats": {
            "gameIdColumn": game_col,
            "uniqueGames": len(games),
            "rowsPerGame": {
                "min": min((len(v) for v in games.values()), default=0),
                "max": max((len(v) for v in games.values()), default=0),
                "median": sorted(len(v) for v in games.values())[len(games) // 2] if games else 0,
            },
            "uniquePlayersPerGame": {
                "min": min(row_counts) if row_counts else 0,
                "max": max(row_counts) if row_counts else 0,
                "distribution": {str(k): v for k, v in sorted(row_count_dist.items())},
            },
            "sampleGameIds": sorted(games.keys())[:10],
        },
        "sampleRows": sample_rows,
        "availableSheets": wb.sheetnames,
    }

    wb.close()
    return profile


def main():
    if len(sys.argv) < 2:
        print(
            "Usage: python scripts/inspect-odi-scorecards.py <path-to.xlsx> [sheet_name] [out.json]",
            file=sys.stderr,
        )
        sys.exit(1)

    path = Path(sys.argv[1])
    sheet_name = sys.argv[2] if len(sys.argv) > 2 and not sys.argv[2].endswith(".json") else None
    out_arg = next((a for a in sys.argv[2:] if a.endswith(".json")), None)
    out_path = Path(out_arg) if out_arg else DEFAULT_OUT

    profile = profile_workbook(path, sheet_name)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(profile, indent=2), encoding="utf-8")

    print(f"Wrote profile to {out_path}")
    print(f"  Rows: {profile['rowCount']}, columns: {profile['columnCount']}")
    print(f"  Games: {profile['gameStats']['uniqueGames']} (column: {profile['gameStats']['gameIdColumn']})")
    print(f"  Inferred keys: {profile['inferredKeys']}")


if __name__ == "__main__":
    main()
