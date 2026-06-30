#!/usr/bin/env python3
import openpyxl
from pathlib import Path

path = Path(
    r"c:\Users\b.carson\Downloads\New Zealand v South Africa 63406779 (1).xlsm"
)
wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
wbv = openpyxl.load_workbook(path, read_only=True, data_only=True)
prep = wb["Prep Work"]
prepv = wbv["Prep Work"]

print("=== Bowler header row 23 (V-AA) ===")
for c in range(22, 28):  # V through AA
    col = openpyxl.utils.get_column_letter(c)
    print(f"{col}23: {prepv[f'{col}23'].value}")

print("\n=== Row 24 bowler inputs formulas ===")
for col in ["V", "W", "X", "Y", "Z", "AA"]:
    addr = f"{col}24"
    print(f"{addr} val={prepv[addr].value}")
    print(f"  formula={prep[addr].value}")

print("\n=== W66, Y66, X66 (T20 man benchmarks inside W63) ===")
for c in ["W66", "Y66", "X66", "Z66", "W63", "Y63", "X63"]:
    print(f"{c}={prepv[c].value}")

# Validate Z calc for row 24-29
w63, y63, x63 = prepv["W63"].value, prepv["Y63"].value, prepv["X63"].value
print(f"\nUsing W63={w63} Y63={y63} X63={x63}")
for row in range(24, 35):
    v, w, x = prepv[f"V{row}"].value, prepv[f"W{row}"].value, prepv[f"X{row}"].value
    z_excel = prepv[f"Z{row}"].value
    if v is None:
        continue
    z_calc = v * ((w63 - w) + y63 * (x - x63)) if v else 0
    print(
        f"row {row} V={v} W={w} X={x} Z_excel={z_excel:.4f} Z_calc={z_calc:.4f} "
        f"diff={z_excel - z_calc:+.4f}"
    )

wb.close()
wbv.close()
