#!/usr/bin/env python3
"""Dump Prep Work B23:F39 and B44:F60 formulas."""
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

path = Path(
    sys.argv[1]
    if len(sys.argv) > 1
    else r"c:\Users\b.carson\Downloads\New Zealand v South Africa 63406779 (1).xlsm"
)

wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
wbv = openpyxl.load_workbook(path, read_only=True, data_only=True)
prep = wb["Prep Work"]
prepv = wbv["Prep Work"]

for name, r1, r2 in [("home", 23, 39), ("away", 44, 60)]:
    print(f"\n=== {name} B{r1}:F{r2} ===")
    for r in range(r1, r2 + 1):
        parts = []
        for c in range(2, 7):
            addr = f"{get_column_letter(c)}{r}"
            f = prep[addr].value
            v = prepv[addr].value
            if f is None and v is None:
                continue
            fs = str(f)[:120] if f else ""
            parts.append(f"{addr}={v!r} f={fs}")
        if parts:
            print(f"R{r}: " + " | ".join(parts))

wb.close()
wbv.close()
