#!/usr/bin/env python3
"""Extract pre-match model outputs and inputs from a cricket trading workbook."""

import json
import sys
from pathlib import Path

import openpyxl


def extract(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    prep = wb["Prep Work"]
    match_info = wb["Match Info"]

    match_id = path.stem.split()[-1].replace("(", "").replace(")", "")

    data = {
        "fixture": {
            "matchId": match_id,
            "homeTeam": match_info["C8"].value,
            "awayTeam": match_info["C9"].value,
            "venue": match_info["C12"].value,
            "format": prep["A1"].value,
            "phase": "pre_match",
            "filename": path.name,
        },
        "matchMarket": {
            "sheet": "Prep Work",
            "marketType": prep["B9"].value,
            "selections": [
                {
                    "team": prep["B10"].value,
                    "probability": prep["C10"].value,
                    "price": prep["D10"].value,
                    "probabilityCell": "C10",
                    "priceCell": "D10",
                },
                {
                    "team": prep["B11"].value,
                    "probability": prep["C11"].value,
                    "price": prep["D11"].value,
                    "probabilityCell": "C11",
                    "priceCell": "D11",
                },
            ],
        },
        "prepInputs": {
            "conditions": {"cell": "D3", "value": prep["D3"].value},
            "battingRating1": {
                "cell": "D4",
                "namedRange": "battingRating1",
                "value": prep["D4"].value,
                "label": prep["B4"].value,
            },
            "battingRating2": {
                "cell": "I4",
                "namedRange": "battingRating2",
                "value": prep["I4"].value,
                "label": prep["G4"].value,
            },
            "bowlingRating1": {
                "cell": "D5",
                "namedRange": "bowlingRating1",
                "value": prep["D5"].value,
                "label": prep["B5"].value,
            },
            "bowlingRating2": {
                "cell": "I5",
                "namedRange": "bowlingRating2",
                "value": prep["I5"].value,
                "label": prep["G5"].value,
            },
            "totalFactor": {"cell": "D6", "value": prep["D6"].value, "team2Cell": "I6", "team2Value": prep["I6"].value},
            "expectedInningsRuns": {
                "team1": {"cell": "E6", "value": prep["E6"].value},
                "team2": {"cell": "J6", "value": prep["J6"].value},
            },
            "parScore": {"cell": "BT3", "value": prep["BT3"].value},
        },
        "sheets": {
            "preMatch": [s for s in wb.sheetnames if s.startswith("PM") or s == "Prep Work"],
            "inPlay": [s for s in wb.sheetnames if s in ("UI", "Scoring", "Pricing", "Input", "Scorecard")],
        },
    }

    wb.close()
    return data


if __name__ == "__main__":
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        r"c:\Users\b.carson\Downloads\New Zealand v South Africa 63406779 (1).xlsm"
    )
    print(json.dumps(extract(workbook_path), indent=2, default=str))
