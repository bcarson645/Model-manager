#!/usr/bin/env python3
"""Extract Prep Work match derivative stats (K2:P18, T2:Z17, team blocks)."""
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import coordinate_to_tuple, get_column_letter

path = Path(
    sys.argv[1]
    if len(sys.argv) > 1
    else r"c:\Users\b.carson\Downloads\New Zealand v South Africa 63406779 (1).xlsm"
)

wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
prep = wb["Prep Work"]


def cell(addr: str):
    row, col = coordinate_to_tuple(addr)
    for r in prep.iter_rows(
        min_row=row, max_row=row, min_col=col, max_col=col, values_only=True
    ):
        return r[0]
    return None


def range_dump(min_row, max_row, min_col, max_col):
    out = {}
    for row in range(min_row, max_row + 1):
        row_data = {}
        for col in range(min_col, max_col + 1):
            v = cell(f"{get_column_letter(col)}{row}")
            if v is not None and v != "":
                row_data[get_column_letter(col)] = v
        if row_data:
            out[str(row)] = row_data
    return out


def sum_col(rows, col_letter):
    total = 0.0
    for row in rows:
        v = cell(f"{col_letter}{row}")
        if isinstance(v, (int, float)):
            total += v
    return total


data = {
    "perInningsHistorical": range_dump(2, 18, 11, 16),
    "matchLevelStats": range_dump(2, 17, 20, 26),
    "teamInnings": {
        "home": {
            "extras": cell("M35"),
            "runOuts": cell("U36"),
            "wickets": cell("U38"),
            "wides": cell("W38"),
        },
        "away": {
            "extras": cell("M56"),
            "runOuts": cell("U57"),
            "wickets": cell("U59"),
            "wides": cell("W59"),
        },
    },
    "limitedOversTotals": {
        "extras": (cell("M35") or 0) + (cell("M56") or 0),
        "runOuts": (cell("U36") or 0) + (cell("U57") or 0),
        "wickets": (cell("U38") or 0) + (cell("U59") or 0),
        "wides": (cell("W38") or 0) + (cell("W59") or 0),
        "ducksBcol": sum_col(range(25, 35), "B") + sum_col(range(46, 56), "B"),
    },
}

print(json.dumps(data, indent=2, default=str))
wb.close()
