#!/usr/bin/env python3
import json
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl


def main():
    path = Path(
        sys.argv[1]
        if len(sys.argv) > 1
        else r"c:\Users\b.carson\Downloads\New Zealand v South Africa 63406779 (1).xlsm"
    )
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    pm = wb["PM Publication"]

    markets = OrderedDict()
    for r in range(1, pm.max_row + 1):
        cat = pm.cell(r, 3).value
        market = pm.cell(r, 4).value
        sel = pm.cell(r, 5).value
        if market and isinstance(market, str) and market not in ("Market1", "Market2"):
            key = (str(cat or ""), str(market))
            if key not in markets:
                markets[key] = {
                    "category": cat,
                    "market": market,
                    "firstRow": r,
                    "selectionCount": 0,
                    "sampleSelections": [],
                }
            markets[key]["selectionCount"] += 1
            markets[key]["lastRow"] = r
            if sel and len(markets[key]["sampleSelections"]) < 2:
                markets[key]["sampleSelections"].append(str(sel)[:50])

    result = list(markets.values())
    print(json.dumps(result, indent=2))
    print(f"\n# unique markets: {len(result)}", file=sys.stderr)
    wb.close()


if __name__ == "__main__":
    main()
