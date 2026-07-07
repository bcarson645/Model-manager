#!/usr/bin/env python3
"""Profile Atlas SRL workbook: SRL sheet, PM Publication, Prep Work."""
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

path = Path(
    sys.argv[1]
    if len(sys.argv) > 1
    else r"c:\Users\b.carson\Downloads\Atlas 196 (3).xlsm"
)


def dump_sheet_preview(wb, name, max_row=80, max_col=20):
    if name not in wb.sheetnames:
        return
    ws = wb[name]
    print(f"\n=== {name} (dims ~{ws.max_row}x{ws.max_column}) ===")
    for r in range(1, min(max_row, ws.max_row or max_row) + 1):
        cells = []
        for c in range(1, min(max_col, ws.max_column or max_col) + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                cells.append(f"{get_column_letter(c)}{r}={v!r}"[:60])
        if cells:
            print(f"R{r}: " + " | ".join(cells[:6]))


def scan_pm_publication(wb):
    ws = wb["PM Publication"]
    print("\n=== PM Publication market rows (A,B,F,G,H,I scan) ===")
    markets = []
    for r in range(1, min(250, ws.max_row or 250) + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        f = ws.cell(r, 6).value
        g = ws.cell(r, 7).value
        h = ws.cell(r, 8).value
        i = ws.cell(r, 9).value
        if a or b or (f is not None) or (g is not None):
            if b or (isinstance(a, (int, float)) and a):
                markets.append(
                    {
                        "row": r,
                        "A": a,
                        "B": str(b)[:80] if b else None,
                        "F": f,
                        "G": g,
                        "H": h,
                        "I": i,
                    }
                )
    print(f"Found {len(markets)} candidate rows")
    for m in markets[:40]:
        print(m)
    if len(markets) > 40:
        print(f"... +{len(markets)-40} more")
    return markets


def scan_srl_sheet(wb):
    ws = wb["SRL"]
    print("\n=== SRL sheet full column A-D scan ===")
    for r in range(1, min(120, ws.max_row or 120) + 1):
        row = []
        for c in range(1, 8):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                row.append(f"{get_column_letter(c)}={v!r}"[:70])
        if row:
            print(f"R{r}: " + " | ".join(row))


wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
print("Workbook:", path.name)
print("Sheets:", len(wb.sheetnames))

scan_srl_sheet(wb)
dump_sheet_preview(wb, "Prep Work", 30, 12)
dump_sheet_preview(wb, "Match Info", 25, 10)

markets = scan_pm_publication(wb)

# Match info fixture
mi = wb["Match Info"]
print("\n=== Match Info key cells ===")
for addr in ["C8", "C9", "C10", "C11", "B3", "C3"]:
    print(f"{addr}={mi[addr].value}")

wb.close()
