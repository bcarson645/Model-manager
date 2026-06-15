#!/usr/bin/env python3
"""Normalize PM Publication markets to base templates (strip team/player names)."""

import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl

TEAM_PATTERNS = [
    r"New Zealand\s*",
    r"South Africa\s*",
    r"NZ\s*",
    r"SA\s*",
]


def normalize_market_name(name: str, category: str) -> str:
    n = name.strip()

    # Player markets: "{Player} - {MarketType}" -> "Player - {MarketType}"
    if category == "Player Market":
        m = re.match(r"^.+\s-\s(.+)$", n)
        if m:
            return f"Player - {m.group(1).strip()}"

    # Team markets: strip team prefix
    for pat in TEAM_PATTERNS:
        n = re.sub(pat, "", n, flags=re.I)

    n = re.sub(r"\s+", " ", n).strip()
    return n or name


def main():
    path = Path(
        sys.argv[1]
        if len(sys.argv) > 1
        else r"c:\Users\b.carson\Downloads\New Zealand v South Africa 63406779 (1).xlsm"
    )
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    pm = wb["PM Publication"]

    templates = OrderedDict()
    for r in range(1, pm.max_row + 1):
        cat = pm.cell(r, 3).value
        market = pm.cell(r, 4).value
        if not market or not isinstance(market, str) or market in ("Market1", "Market2"):
            continue
        cat_s = str(cat or "Unknown")
        base = normalize_market_name(market, cat_s)
        key = (cat_s, base)
        if key not in templates:
            templates[key] = {
                "category": cat_s,
                "marketTemplate": base,
                "exampleNames": [],
                "firstRow": r,
                "lastRow": r,
                "selectionCount": 0,
                "instanceCount": 0,
            }
        t = templates[key]
        t["instanceCount"] += 1
        t["selectionCount"] += 1
        t["lastRow"] = r
        if market not in t["exampleNames"] and len(t["exampleNames"]) < 3:
            t["exampleNames"].append(market)

    result = list(templates.values())
    out = Path(__file__).resolve().parent.parent / "lib" / "workbooks" / "pm-market-templates.json"
    out.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(f"Wrote {len(result)} templates to {out}", file=sys.stderr)
    wb.close()


if __name__ == "__main__":
    main()
